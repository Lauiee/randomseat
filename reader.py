"""
엑셀 파일 읽기: 좌석 배치(시트1), 참가자 명단(시트2)
"""
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class Seat:
    """좌석 하나"""
    cell_ref: str  # 예: "A4"
    row: int
    col: int


@dataclass
class Participant:
    """참가자"""
    name: str
    gisu: str  # 기수


def _is_solid_fill(cell) -> bool:
    """solid 패턴 채우기 = 통로/비좌석"""
    fill = getattr(cell, "fill", None)
    if not fill:
        return False
    return getattr(fill, "patternType", None) == "solid"


def read_seats(
    excel_path: str | Path,
    seat_marker: str = "O",
) -> tuple[list[Seat], dict[str, str]]:
    """
    시트1에서 좌석 추출.
    - solid fill(회색) = 비좌석
    - 좌석 영역 내 흰색 셀 = 좌석 (O/이름/빈칸 모두)
    - O 또는 빈칸 = 배정 가능 좌석
    - 이름이 있으면 = 고정 좌석
    반환: (배정가능 좌석 목록, 고정 배치 {이름: 셀ref})
    """
    wb = load_workbook(excel_path)
    ws = wb.worksheets[0]
    marker = str(seat_marker).strip().upper()
    min_r = min_c = 999
    max_r = max_c = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 20, min_col=1, max_col=ws.max_column or 20):
        for cell in row:
            if _is_solid_fill(cell):
                continue
            val = cell.value
            s = str(val).strip() if val else ""
            if "교단" in s:
                continue
            if s.upper() == marker or (len(s) >= 1 and s != "O"):  # O 또는 이름 있는 셀로만 영역 판단
                min_r, max_r = min(min_r, cell.row), max(max_r, cell.row)
                min_c, max_c = min(min_c, cell.column), max(max_c, cell.column)

    seats: list[Seat] = []
    fixed: dict[str, str] = {}
    for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
        for cell in row:
            if _is_solid_fill(cell):
                continue
            val = cell.value
            s = str(val).strip() if val else ""
            if "교단" in s:
                continue
            if s.upper() == marker:
                # O 있는 좌석만 배정 대상
                seats.append(Seat(cell_ref=cell.coordinate, row=cell.row, col=cell.column))
            elif len(s) >= 1:
                fixed[s] = cell.coordinate

    wb.close()
    return seats, fixed


def read_participants(excel_path: str | Path, name_col: str = "이름", gisu_col: str = "기수") -> list[Participant]:
    """
    시트2에서 참가자 명단 추출.
    첫 행을 헤더로 보고, name_col, gisu_col 컬럼 사용.
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.worksheets[1]  # 시트2
    header = None
    name_idx = gisu_idx = -1
    participants: list[Participant] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row or 200, values_only=True), start=1):
        cells = list(row)
        if row_idx == 1:
            header = [str(c).strip() if c else "" for c in cells]
            for i, h in enumerate(header):
                if h in (name_col, "이름", "name", "Name"):
                    name_idx = i
                if h in (gisu_col, "기수", "gisu", "Gisu", "batch", "Batch"):
                    gisu_idx = i
            if name_idx < 0:
                name_idx = 0  # 기본값
            if gisu_idx < 0:
                gisu_idx = 1 if len(header) > 1 else 0
            continue

        if not cells:
            continue
        name = str(cells[name_idx] or "").strip()
        if not name:
            continue
        gisu = str(cells[gisu_idx] or "").strip()
        participants.append(Participant(name=name, gisu=gisu))

    wb.close()
    return participants
