"""
배치 결과를 이미지로 출력
"""
import re
from pathlib import Path

from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


def _parse_cell_ref(ref: str) -> tuple[int, int]:
    """A1 -> (1, 1), B3 -> (3, 2) (row, col)"""
    m = re.match(r"([A-Z]+)(\d+)$", ref, re.I)
    if not m:
        return 0, 0
    col = column_index_from_string(m.group(1))
    row = int(m.group(2))
    return row, col


def generate_image(
    assignments: dict[str, str],
    excel_path: str | Path,
    output_path: str | Path,
    cell_size: int = 70,
    font_size: int = 14,
) -> None:
    """
    좌석 배치 시트와 동일한 구조로 이름이 채워진 이미지 생성.
    assignments: { "이름": "A4", ... }
    """
    wb = load_workbook(excel_path)
    ws = wb.worksheets[0]
    max_row = ws.max_row or 13
    max_col = ws.max_column or 11

    # cell_ref -> name 매핑
    ref_to_name = {v: k for k, v in assignments.items()}

    def is_solid_fill(c):
        fill = getattr(c, "fill", None)
        return fill and getattr(fill, "patternType", None) == "solid"

    def is_light_gray_fill(c) -> bool:
        """연한 회색(미사용 좌석) vs 진한 회색(통로) 구분"""
        try:
            fill = getattr(c, "fill", None)
            if not fill or getattr(fill, "patternType", None) != "solid":
                return False
            rgb = getattr(fill.fgColor, "rgb", None)
            if not rgb or not isinstance(rgb, str) or len(rgb) < 6:
                return False
            h = rgb[2:] if len(rgb) == 8 else rgb
            if len(h) == 6:
                r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
                return (r + g + b) / 3 >= 200
        except Exception:
            pass
        return False

    # 좌석 영역: O/이름이 있는 범위 (교단 제외)
    min_r = min_c = 999
    max_r = max_c = 0
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if is_solid_fill(cell):
                continue
            val = cell.value
            s = str(val).strip() if val else ""
            if "교단" in s:
                continue
            if s.upper() == "O" or (len(s) >= 1 and "교단" not in s):  # O 또는 이름
                min_r, max_r = min(min_r, cell.row), max(max_r, cell.row)
                min_c, max_c = min(min_c, cell.column), max(max_c, cell.column)

    seat_refs = set()
    podium_refs: dict[str, str] = {}
    merged_podium: list[tuple[int, int, int, int, str]] = []
    merged_skip: set[tuple[int, int]] = set()
    if min_r > max_r:
        min_r, max_r = 1, max_row
        min_c, max_c = 1, max_col

    # 교단: 시트 전체에서 검색 (좌석 영역 밖에 있을 수 있음)
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            s = str(cell.value or "").strip()
            if "교단" in s:
                podium_refs[cell.coordinate] = s

    # 좌석: 영역 확장(min_r~max_row) 후 solid 아닌 셀 + 연한 회색(미사용 좌석)
    for row in ws.iter_rows(min_row=min_r, max_row=max_row, min_col=min_c, max_col=max_col):
        for cell in row:
            if "교단" in str(cell.value or ""):
                continue
            if is_solid_fill(cell):
                if is_light_gray_fill(cell):
                    seat_refs.add(cell.coordinate)
                continue
            seat_refs.add(cell.coordinate)

    # 합병된 셀 처리 (교단 등)
    for mr in ws.merged_cells.ranges:
        ref_str = str(mr)
        if ":" in ref_str:
            start, end = ref_str.split(":")
            r1, c1 = _parse_cell_ref(start.strip())
            r2, c2 = _parse_cell_ref(end.strip())
            top_left = f"{get_column_letter(c1)}{r1}"
            txt = podium_refs.pop(top_left, None)
            if txt:
                merged_podium.append((r1, c1, r2, c2, txt))
                for r in range(r1, r2 + 1):
                    for c in range(c1, c2 + 1):
                        merged_skip.add((r, c))

    # 아래에서 3번째 행과 교단 사이 = 회색(비좌석). seat_refs에 추가하지 않음.
    wb.close()

    # 이미지 크기: 헤더 없음
    header_h = 0
    w = max_col * cell_size
    h = header_h + max_row * cell_size

    img = Image.new("RGB", (w, h), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)

    try:
        font = ImageFont.truetype("/System/Library/Fonts/AppleSDGothicNeo.ttc", font_size)
    except OSError:
        font = ImageFont.load_default()

    def draw_rect(x1: int, y1: int, x2: int, y2: int, fill_color: tuple, outline_color: tuple) -> None:
        draw.rectangle([x1, y1, x2, y2], outline=outline_color, fill=fill_color)

    # 합병된 교단 영역 먼저 그리기
    try:
        podium_font = ImageFont.truetype("/System/Library/Fonts/AppleSDGothicNeo.ttc", font_size + 4)
    except OSError:
        podium_font = font
    for r1, c1, r2, c2, _ in merged_podium:
        x1 = (c1 - 1) * cell_size
        y1 = header_h + (r1 - 1) * cell_size
        x2 = c2 * cell_size
        y2 = header_h + r2 * cell_size
        draw_rect(x1, y1, x2, y2, (170, 170, 170), (150, 150, 150))
        txt = "교단"
        bbox = draw.textbbox((0, 0), txt, font=podium_font)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        tx = x1 + (x2 - x1 - tw) // 2
        ty = y1 + (y2 - y1 - th) // 2
        draw.text((tx, ty), txt, fill=(40, 40, 40), font=podium_font)

    # 그리드 및 셀 내용
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in merged_skip:
                continue  # 합병 셀 내부는 이미 그렸음
            cell_ref = f"{get_column_letter(c)}{r}"
            x = (c - 1) * cell_size
            y = header_h + (r - 1) * cell_size

            is_seat = cell_ref in seat_refs
            is_podium = cell_ref in podium_refs
            has_person = bool(ref_to_name.get(cell_ref, ""))
            if is_seat and has_person:
                fill = (255, 255, 255)  # 사용 중 좌석: 흰색
                outline = (180, 180, 180)
            elif is_seat:
                fill = (248, 252, 255)  # 비어 있는 좌석: 연한 파란 톤
                outline = (200, 215, 230)
            elif is_podium:
                fill = (170, 170, 170)  # 교단: 진한 회색
                outline = (150, 150, 150)
            else:
                fill = (210, 210, 210)  # 좌석 아님(통로 등): 중간 회색
                outline = (190, 190, 190)

            draw.rectangle([x, y, x + cell_size, y + cell_size], outline=outline, fill=fill)

            name = ref_to_name.get(cell_ref, "") if is_seat else ""
            if is_podium:
                txt = podium_refs[cell_ref]
                bbox = draw.textbbox((0, 0), txt, font=font)
                tw = bbox[2] - bbox[0]
                th = bbox[3] - bbox[1]
                tx = x + (cell_size - tw) // 2
                ty = y + (cell_size - th) // 2
                draw.text((tx, ty), txt, fill=(80, 80, 80), font=font)
            elif is_seat and name:
                # 긴 이름은 줄임
                txt = name[:6] + "…" if len(name) > 6 else name
                bbox = draw.textbbox((0, 0), txt, font=font)
                tw = bbox[2] - bbox[0]
                th = bbox[3] - bbox[1]
                tx = x + (cell_size - tw) // 2
                ty = y + (cell_size - th) // 2
                draw.text((tx, ty), txt, fill=(0, 0, 0), font=font)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    img.save(out)
    print(f"이미지 저장: {out}")
