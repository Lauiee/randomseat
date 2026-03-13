"""
샘플 엑셀 파일 생성 (테스트용)
실행: python create_sample.py
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 시트1: 좌석 배치 (O = 좌석, 비어있거나 회색 = 통로/비좌석)
# 사용자 제공 레이아웃 참고 (41석, 3블록)
LAYOUT = [
    # 행4~10 정도, A-K열
    # A,B,C / D(통로) / E,F,G / H(통로) / I,J,K
    # 간단히 4x11 그리드에 O 배치
    ["", "", "", "", "", "", "", "", "", "", ""],  # 1
    ["", "", "", "", "", "", "", "", "", "", ""],  # 2
    ["", "", "", "", "", "", "", "", "", "", ""],  # 3
    ["O", "", "O", "", "O", "", "O", "", "O", "", "O"],  # 4
    ["O", "", "O", "", "O", "", "O", "", "O", "", "O"],  # 5
    ["O", "", "O", "", "O", "", "O", "", "O", "", "O"],  # 6
    ["O", "", "O", "", "O", "", "O", "", "O", "", "O"],  # 7
    ["O", "", "O", "", "O", "O", "O", "", "O", "", "O"],  # 8
    ["", "", "O", "", "O", "O", "O", "", "O", "", ""],  # 9
    ["", "", "O", "", "O", "O", "O", "", "O", "", ""],  # 10
    ["", "", "", "", "", "", "", "", "", "", ""],  # 11
    ["", "", "", "", "", "", "", "", "", "", ""],  # 12
    ["", "", "", "", "", "교단", "", "", "", "", ""],  # 13
]


def main():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "좌석배치"

    for r, row_data in enumerate(LAYOUT, start=1):
        for c, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=r, column=c, value=val if val else None)
            if val == "교단":
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    ws2 = wb.create_sheet("참가자명단", 1)
    ws2.append(["이름", "기수"])
    # 샘플 30명 (기수 1~3)
    for i in range(1, 31):
        name = f"참가자{i}"
        gisu = str((i - 1) % 3 + 1)  # 1,2,3 반복
        ws2.append([name, gisu])

    out = "sample.xlsx"
    wb.save(out)
    print(f"생성됨: {out}")


if __name__ == "__main__":
    main()
